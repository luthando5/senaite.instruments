# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.INSTRUMENTS.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2019 by it's authors.
# Some rights reserved, see README and LICENSE.
import csv
import json
import types
import traceback
from cStringIO import StringIO
from mimetypes import guess_type
from openpyxl import load_workbook
from os.path import abspath
from os.path import basename
from os.path import splitext
from re import subn
from xlrd import open_workbook

from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface, IInstrumentImportInterface
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter)
from senaite.core.exportimport.instruments.resultsimport import (
    InstrumentResultsFileParser)

from bika.lims import api
from bika.lims import bikaMessageFactory as _
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.core.catalog import ANALYSIS_CATALOG, SENAITE_CATALOG
from senaite.instruments.instrument import FileStub
from senaite.instruments.instrument import SheetNotFound
from zope.interface import implements
from zope.publisher.browser import FileUpload

field_interim_map = {
    "Formula": "formula",
    "Concentration": "concentration",
    "Z": "z",
    "Status": "status",
    "Line 1": "line_1",
    "Net int.": "net_int",
    "LLD": "lld",
    "Stat. error": "stat_error",
    "Analyzed layer": "analyzed_layer",
    "Bound %": "bound_pct",
}


class SampleNotFound(Exception):
    pass


class MultipleAnalysesFound(Exception):
    pass


class AnalysisNotFound(Exception):
    pass


class S8TigerParser(InstrumentResultsFileParser):
    ar = None

    def __init__(self, infile, worksheet=None, encoding=None, delimiter=None):
        self.delimiter = delimiter if delimiter else ","
        self.encoding = encoding
        self.ar = None
        self.analyses = None
        self.worksheet = worksheet if worksheet else 0
        self.infile = infile
        self.csv_data = None
        self.sample_id = None
        mimetype, encoding = guess_type(self.infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def xls_to_csv(self, infile, worksheet=0, delimiter=","):
        """
        Convert xlsx to easier format first, since we want to use the
        convenience of the CSV library

        """

        def find_sheet(wb, worksheet):
            for sheet in wb.sheets():
                if sheet.name == worksheet:
                    return sheet

        wb = open_workbook(file_contents=infile.read())
        sheet = wb.sheets()[worksheet]

        buffer = StringIO()

        # extract all rows
        for row in sheet.get_rows():
            line = []
            for cell in row:
                value = cell.value
                if type(value) in types.StringTypes:
                    value = value.encode("utf8")
                if value is None:
                    value = ""
                line.append(str(value))
            print >> buffer, delimiter.join(line)
        buffer.seek(0)
        return buffer

    def xlsx_to_csv(self, infile, worksheet=None, delimiter=","):
        worksheet = worksheet if worksheet else 0
        wb = load_workbook(filename=infile)
        if worksheet in wb.sheetnames:
            sheet = wb[worksheet]
        else:
            try:
                index = int(worksheet)
                sheet = wb.worksheets[index]
            except (ValueError, TypeError, IndexError):
                raise SheetNotFound

        buffer = StringIO()
        for row in sheet.rows:
            line = []
            for cell in row:
                new_val = ''
                if cell.number_format == "0.00%":
                    new_val = '{}%'.format(cell.value * 100)
                cellval = new_val if new_val else cell.value

                value = "" if cellval is None else str(cellval).encode("utf8")
                if "\n" in value:  # fixme multi-line cell gives only first line
                    value = value.split("\n")[0]
                line.append(value.strip())
            if not any(line):
                continue
            buffer.write(delimiter.join(line) + "\n")
        buffer.seek(0)
        return buffer

    def parse(self):
        order = []
        ext = splitext(self.infile.filename.lower())[-1]
        if ext == ".xlsx":
            order = (self.xlsx_to_csv, self.xls_to_csv)
        elif ext == ".xls":
            order = (self.xls_to_csv, self.xlsx_to_csv)
        elif ext == ".csv":
            self.csv_data = self.infile
        if order:
            for importer in order:
                try:
                    self.csv_data = importer(
                        infile=self.infile,
                        worksheet=self.worksheet,
                        delimiter=self.delimiter,
                    )
                    break
                except SheetNotFound:
                    self.err("Sheet not found in workbook: %s" % self.worksheet)
                    return -1
                except Exception as e:  # noqa
                    pass
            else:
                self.warn("Can't parse input file as XLS, XLSX, or CSV.")
                return -1
        stub = FileStub(file=self.csv_data, name=str(self.infile.filename))
        self.csv_data = FileUpload(stub)

        try:
            sample_id, ext = splitext(basename(self.infile.filename))
            portal_type = self.get_portal_type(sample_id)
            # Check for sample, duplicates and reference analysis(QC)
            if not portal_type:
                # maybe we need to chop of it's -9digit suffix
                sample_id = "-".join(sample_id.split("-")[:-1])
                portal_type = self.get_portal_type(sample_id)
                if not portal_type:
                    # or we are out of luck
                    msg = "Can't find sample for " + self.infile.filename
                    self.warn(msg)
                    return -1

            self.sample_id = sample_id
        except Exception as e:
            self.err(repr(e))
            return False
        lines = self.csv_data.readlines()
        reader = csv.DictReader(lines)
        if portal_type == "AnalysisRequest":
            for row in reader:
                self.parse_ar_row(sample_id, reader.line_num, row)

        elif portal_type in ["DuplicateAnalysis", "ReferenceAnalysis"]:
            for row in reader:
                self.parse_duplicate_row(sample_id, reader.line_num, row)

        elif portal_type == "ReferenceSample":
            for row in reader:
                self.parse_reference_sample_row(sample_id, reader.line_num, row)
        return 0

    def get_portal_type(self, sample_id):
        portal_type = None
        if self.is_sample(sample_id):
            ar = self.get_ar(sample_id)
            self.ar = ar
            self.analyses = self.get_analyses(ar)
            portal_type = ar.portal_type
        elif self.is_analysis_group_id(sample_id):
            portal_type = "DuplicateAnalysis"
        elif self.is_reference_sample(sample_id):
            portal_type = "ReferenceSample"
        return portal_type

    def parse_row(self, row_nr, row, keyword):
        parsed = {field_interim_map.get(k, ""): v for k, v in row.items()}
        # Concentration can be PPM or PCT as it likes, I'll save both.
        concentration = parsed["concentration"]
        try:
            val = float(subn(r'[^.\d]', '', str(concentration))[0])
        except (TypeError, ValueError, IndexError):
            self.warn(
                msg="Can't extract numerical value from `concentration`",
                numline=row_nr,
                line=str(row),
            )
            parsed["reading"] = ""
            return 0
        else:
            if "ppm" in concentration.lower():
                parsed["reading"] = val * 0.0001
            elif "%" in concentration:
                parsed["reading"] = val
            else:
                self.warn(
                    "Can't decide if reading units are PPM or %",
                    numline=row_nr,
                    line=str(row),
                )
                return 0

        parsed.update({"DefaultResult": "reading"})

        self._addRawResult(self.sample_id, {keyword: parsed})
        return 0

    def parse_ar_row(self, sample_id, row_nr, row):
        ar = self.get_ar(sample_id)
        # convert row to use interim field names
        parsed = {field_interim_map.get(k, ""): v for k, v in row.items()}

        formula = parsed.get("formula")
        kw = subn(r'[^\w\d\-_]*', '', formula)[0]
        try:
            analysis = self.get_analysis(ar, kw)
            if not analysis:
                return 0
            keyword = analysis.getKeyword
        except Exception as e:
            self.warn(
                msg="Error getting analysis for '${kw}': ${e}",
                mapping={"kw": kw, "e": repr(e)},
                numline=row_nr,
                line=str(row),
            )
            return
        return self.parse_row(row_nr, row, keyword)

    def parse_duplicate_row(self, sample_id, row_nr, row):
        # convert row to use interim field names
        parsed = {field_interim_map.get(k, ""): v for k, v in row.items()}
        try:
            formula = parsed.get("formula")
            kw = subn(r'[^\w\d\-_]*', '', formula)[0]
            keyword = self.getDuplicateKeyord(sample_id, kw)
        except Exception as e:
            self.warn(
                msg="Error getting analysis for '${kw}': ${e}",
                mapping={"kw": kw, "e": repr(e)},
                numline=row_nr,
                line=str(row),
            )
            return
        return self.parse_row(row_nr, row, keyword)

    def getDuplicateKeyord(self, sample_id, kw):
        analysis = self.get_duplicate_or_qc_analysis(sample_id, kw)
        return analysis.getKeyword

    def parse_reference_sample_row(self, sample_id, row_nr, row):
        # convert row to use interim field names
        parsed = {field_interim_map.get(k, ""): v for k, v in row.items()}
        try:
            formula = parsed.get("formula")
            kw = subn(r'[^\w\d\-_]*', '', formula)[0]
            keyword = self.getReferenceSampleKeyword(sample_id, kw)
        except Exception as e:
            self.warn(
                msg="Error getting analysis for '${kw}': ${e}",
                mapping={"kw": kw, "e": repr(e)},
                numline=row_nr,
                line=str(row),
            )
            return
        return self.parse_row(row_nr, row, keyword)

    def getReferenceSampleKeyword(self, sample_id, kw):
        sample_reference = self.get_reference_sample(sample_id, kw)
        analysis = self.get_reference_sample_analysis(sample_reference, kw)
        return analysis.getKeyword()

    @staticmethod
    def get_duplicate_or_qc_analysis(analysis_id, kw):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types, getReferenceAnalysesGroupID=analysis_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        analyses = dict((a.getKeyword, a) for a in brains)
        brains = [v for k, v in analyses.items() if k.startswith(kw)]
        if len(brains) < 1:
            msg = ("No analysis found matching Keyword '${kw}'",)
            raise AnalysisNotFound(msg, kw=kw)
        if len(brains) > 1:
            msg = ("Multiple brains found matching Keyword '${kw}'",)
            raise MultipleAnalysesFound(msg, kw=kw)
        return brains[0]

    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        try:
            return api.get_object(brains[0])
        except IndexError:
            pass

    @staticmethod
    def is_sample(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return True if brains else False

    @staticmethod
    def get_analyses(ar):
        analyses = ar.getAnalyses()
        return dict((a.getKeyword, a) for a in analyses)

    def get_analysis(self, ar, kw):
        analyses = self.get_analyses(ar)
        analyses = [v for k, v in analyses.items() if k.startswith(kw)]
        if len(analyses) < 1:
            self.log('No analysis found matching keyword "${kw}"', mapping=dict(kw=kw))
            return None
        if len(analyses) > 1:
            self.warn(
                'Multiple analyses found matching Keyword "${kw}"', mapping=dict(kw=kw)
            )
            return None
        return analyses[0]

    @staticmethod
    def is_analysis_group_id(analysis_group_id):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types, getReferenceAnalysesGroupID=analysis_group_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        return True if brains else False

    @staticmethod
    def is_reference_sample(reference_sample_id):
        query = dict(
            portal_type="ReferenceSample", getId=reference_sample_id
        )
        brains = api.search(query, SENAITE_CATALOG)
        return True if brains else False

    @staticmethod
    def get_reference_sample(reference_sample_id, kw):
        query = dict(
            portal_type="ReferenceSample", getId=reference_sample_id
        )
        brains = api.search(query, SENAITE_CATALOG)
        if len(brains) < 1:
            msg = ("No reference sample found matching Keyword '${kw}'",)
            raise AnalysisNotFound(msg, kw=kw)
        if len(brains) > 1:
            msg = ("Multiple brains found matching Keyword '{}'".format(kw))
            raise MultipleAnalysesFound(msg)
        return brains[0]


class importer(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Bruker S8 Tiger"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        worksheet = request.form.get("worksheet", 0)
        parser = S8TigerParser(infile, worksheet=worksheet)
        if parser:

            status = ["sample_received", "attachment_due", "to_be_verified"]
            if artoapply == "received":
                status = ["sample_received"]
            elif artoapply == "received_tobeverified":
                status = ["sample_received", "attachment_due", "to_be_verified"]

            over = [False, False]
            if override == "nooverride":
                over = [False, False]
            elif override == "override":
                over = [True, False]
            elif override == "overrideempty":
                over = [True, True]

            importer = AnalysisResultsImporter(
                parser=parser,
                context=context,
                allowed_ar_states=status,
                allowed_analysis_states=None,
                override=over,
                instrument_uid=instrument,
            )

            try:
                importer.process()
                errors = importer.errors
                logs = importer.logs
                warns = importer.warns
            except Exception as e:
                errors.extend([repr(e), traceback.format_exc()])

        results = {"errors": errors, "log": logs, "warns": warns}

        return json.dumps(results)
