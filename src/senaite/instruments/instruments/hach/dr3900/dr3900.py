# -*- coding: utf-8 -*-
#
# This file is part of SENAITE.INSTRUMENTS.
#
# SENAITE.CORE is free software: you can redistribute it and/or modify it under
# the terms of the GNU General Public License as published by the Free Software
# Foundation, version 2.
#
# This program is distributed in the hope that it will be useful, but WITHOUT
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS
# FOR A PARTICULAR PURPOSE. See the GNU General Public License for more
# details.
#
# You should have received a copy of the GNU General Public License along with
# this program; if not, write to the Free Software Foundation, Inc., 51
# Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# Copyright 2018-2019 by it's authors.
# Some rights reserved, see README and LICENSE.

# import csv # Unused Code
import json
import types
import traceback
from cStringIO import StringIO
from mimetypes import guess_type
from openpyxl import load_workbook
from os.path import abspath
from os.path import splitext
from xlrd import open_workbook

from senaite.core.exportimport.instruments import (
    IInstrumentAutoImportInterface, IInstrumentImportInterface
)
from senaite.core.exportimport.instruments.resultsimport import (
    AnalysisResultsImporter)
from senaite.core.exportimport.instruments.resultsimport import (
    InstrumentResultsFileParser)

from bika.lims import api
from bika.lims import bikaMessageFactory as _
from bika.lims.catalog import CATALOG_ANALYSIS_REQUEST_LISTING
from senaite.core.catalog import ANALYSIS_CATALOG
from senaite.instruments.instrument import FileStub
from senaite.instruments.instrument import SheetNotFound
from zope.interface import implements
from zope.publisher.browser import FileUpload

# Unused Code
# field_interim_map = {
#     "Formula": "formula",
#     "Concentration": "concentration",
#     "Z": "z",
#     "Status": "status",
#     "Line 1": "line_1",
#     "Net int.": "net_int",
#     "LLD": "lld",
#     "Stat. error": "stat_error",
#     "Analyzed layer": "analyzed_layer",
#     "Bound %": "bound_pct",
# }


class SampleNotFound(Exception):
    pass


class MultipleAnalysesFound(Exception):
    pass


class AnalysisNotFound(Exception):
    pass


class DR3900Parser(InstrumentResultsFileParser):
    ar = None

    def __init__(self, infile, worksheet=None, encoding=None, delimiter=None):
        self.delimiter = delimiter if delimiter else ","
        self.encoding = encoding
        self.ar = None
        self.analyses = None
        self.worksheet = worksheet if worksheet else 0
        self.infile = infile
        self.csv_data = None
        self.sample_id = None
        self.processed_samples_class = []
        mimetype, encoding = guess_type(self.infile.filename)
        InstrumentResultsFileParser.__init__(self, infile, mimetype)

    def xls_to_csv(self, infile, worksheet=0, delimiter=","):
        """
        Convert xlsx to easier format first, since we want to use the
        convenience of the CSV library
        """

        def find_sheet(wb, worksheet):
            for sheet in wb.sheets():
                if sheet.name == worksheet:
                    return sheet

        wb = open_workbook(file_contents=infile.read())
        sheet = wb.sheets()[worksheet]

        buffer = StringIO()

        # extract all rows
        for row in sheet.get_rows():
            line = []
            for cell in row:
                value = cell.value
                if type(value) in types.StringTypes:
                    value = value.encode("utf8")
                if value is None:
                    value = ""
                line.append(str(value))
            print >> buffer, delimiter.join(line)
        buffer.seek(0)
        return buffer

    def xlsx_to_csv(self, infile, worksheet=None, delimiter=","):
        worksheet = worksheet if worksheet else 0
        wb = load_workbook(filename=infile)
        if worksheet in wb.sheetnames:
            sheet = wb[worksheet]
        else:
            try:
                index = int(worksheet)
                sheet = wb.worksheets[index]
            except (ValueError, TypeError, IndexError):
                raise SheetNotFound

        buffer = StringIO()

        for row in sheet.rows:
            line = []
            for cell in row:
                new_val = ''
                if cell.number_format == "0.00%":
                    new_val = '{}%'.format(cell.value * 100)
                cellval = new_val if new_val else cell.value
                if (isinstance(cellval, (int, long, float))):
                    value = "" if cellval is None else str(cellval).encode("utf8")
                else:
                    value = "" if cellval is None else cellval.encode("utf8")
                if "\n" in value:
                    value = value.split("\n")[0]
                line.append(value.strip())
            if not any(line):
                continue
            buffer.write(delimiter.join(line) + "\n")
        buffer.seek(0)
        return buffer


    def parse(self):
        order = []
        ext = splitext(self.infile.filename.lower())[-1]
        if ext == ".xlsx":
            order = (self.xlsx_to_csv, self.xls_to_csv)
        elif ext == ".xls":
            order = (self.xls_to_csv, self.xlsx_to_csv)
        elif ext == ".csv" or ".prn":
            self.csv_data = self.infile
        if order:
            for importer in order:
                try:
                    self.csv_data = importer(
                        infile=self.infile,
                        worksheet=self.worksheet,
                        delimiter=self.delimiter,
                    )
                    break
                except SheetNotFound:
                    self.err("Sheet not found in workbook: %s" % self.worksheet)
                    return -1
                except Exception as e:
                    pass
            else:
                self.warn("Can't parse input file as XLS, XLSX, or CSV.")
                return -1
        stub = FileStub(file=self.csv_data, name=str(self.infile.filename))
        self.csv_data = FileUpload(stub)

        lines_with_parentheses = self.csv_data.readlines()
        lines = [i.replace('"','').replace('\r\n','') for i in lines_with_parentheses]

        analysis_round = 0
        sample_service = []
        for row_nr, row in enumerate(lines): #This whole part can be a function
            split_row = row.split(",")
            if 'M\xc3\xa9thode:' in split_row[0] or 'M\xe9thode:' in split_row[0]:
                analysis_round = analysis_round + 1
            if 'M\xc3\xa9thodes' in split_row[0] or 'M\xe9thodes' in split_row[0]:
                #Here we determine how many rounds there are in the sheet (Max = 3)
                if split_row[1]:
                    sample_service.append(split_row[1])
                if len(split_row) > 2 and split_row[2]:
                    sample_service.append(split_row[2])
                if len(split_row) > 3 and split_row[3]:
                    sample_service.append(split_row[3])
            if analysis_round > 0 and split_row[0] and len(split_row)>2 and split_row[1]: #How long is split_row of theres an empty cell inbetween?
                #If we are past the headerlines and the first and second columns entries (of that row) are non empty
                self.parse_row(row_nr, split_row,sample_service[analysis_round-1],analysis_round)
        return 0


    def parse_row(self, row_nr, row,sample_service,analysis_round):
        #Try to restructure parse row suc
        parsed = {}
        #Here we check whether this sample ID has been processed already
        if {row[0]:sample_service} in self.processed_samples_class:
            msg = ("Multiple results for Sample '{}' with sample service '{}' found. Not imported".format(row[0],sample_service))
            raise MultipleAnalysesFound(msg)
        if self.is_sample(row[0]):
            sample = self.get_ar(row[0])
        else:
            #Updating the Reference analyses
            sample = self.get_duplicate_or_qc(row[0],sample_service)# change to qc or reference
            if sample:# Don't have to check for sample as an error will be thrown already
                keyword = sample.getKeyword
                self.processed_samples_class.append({row[0]:sample_service})
                parsed["Reading"] = float(row[1])
                parsed["Factor"] = float(row[8])
                parsed.update({"DefaultResult": "Reading"})
                self._addRawResult(row[0], {keyword: parsed})
                return 0
            else:
                return 0
        # Updating the analysis requests
        analyses = sample.getAnalyses()
        for analysis in analyses: #Use getAnalysis instead of getAnalyses using the keyword is the distinguisher
            if sample_service == analysis.getKeyword:
                keyword = analysis.getKeyword
                if row[1] == 'OVER':
                    if analysis_round == 3:
                        #If in the third analysis_round [Reading] = OVER then the value 999999 is assigned.
                        self.processed_samples_class.append({row[0]:sample_service})
                        parsed["Reading"] = float(999999)
                        parsed["Factor"] = float(1)
                        parsed.update({"DefaultResult": "Reading"})
                        self._addRawResult(row[0], {keyword: parsed})
                    #If not in the 3rd analysis_round and Reading = OVER, we don't update the Reading
                    return
                self.processed_samples_class.append({row[0]:sample_service})
                parsed["Reading"] = float(row[1])
                parsed["Factor"] = float(row[8])
                parsed.update({"DefaultResult": "Reading"})
                self._addRawResult(row[0], {keyword: parsed})
                #Avoid repetition
        return 


    @staticmethod
    def get_ar(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        try:
            return api.get_object(brains[0])
        except IndexError:
            pass
    
    @staticmethod
    def is_sample(sample_id):
        query = dict(portal_type="AnalysisRequest", getId=sample_id)
        brains = api.search(query, CATALOG_ANALYSIS_REQUEST_LISTING)
        return True if brains else False


    @staticmethod
    def get_duplicate_or_qc(analysis_id,sample_service,):
        portal_types = ["DuplicateAnalysis", "ReferenceAnalysis"]
        query = dict(
            portal_type=portal_types, getReferenceAnalysesGroupID=analysis_id
        )
        brains = api.search(query, ANALYSIS_CATALOG)
        analyses = dict((a.getKeyword, a) for a in brains)
        brains = [v for k, v in analyses.items() if k.startswith(sample_service)]
        if len(brains) < 1:
            msg = ("No analysis found matching Keyword {}".format(sample_service))
            raise AnalysisNotFound(msg)
        if len(brains) > 1:
            msg = ("Multiple brains found matching Keyword {}".format(sample_service))
            raise MultipleAnalysesFound(msg)
        return brains[0]


    @staticmethod
    def get_analyses(ar):
        analyses = ar.getAnalyses()
        return dict((a.getKeyword, a) for a in analyses)

    def get_analysis(self, ar, kw):
        analyses = self.get_analyses(ar)
        analyses = [v for k, v in analyses.items() if k.startswith(kw)]
        if len(analyses) < 1:
            self.log('No analysis found matching keyword "${kw}"', mapping=dict(kw=kw))
            return None
        if len(analyses) > 1:
            self.warn(
                'Multiple analyses found matching Keyword "${kw}"', mapping=dict(kw=kw)
            )
            return None
        return analyses[0]


class dr3900import(object):
    implements(IInstrumentImportInterface, IInstrumentAutoImportInterface)
    title = "Hach DR3900"
    __file__ = abspath(__file__)  # noqa

    def __init__(self, context):
        self.context = context
        self.request = None

    @staticmethod
    def Import(context, request):
        errors = []
        logs = []
        warns = []

        infile = request.form["instrument_results_file"]
        if not hasattr(infile, "filename"):
            errors.append(_("No file selected"))

        artoapply = request.form["artoapply"]
        override = request.form["results_override"]
        instrument = request.form.get("instrument", None)
        worksheet = request.form.get("worksheet", 0)
        parser = DR3900Parser(infile, worksheet=worksheet)
        if parser:

            status = ["sample_received", "attachment_due", "to_be_verified"]
            if artoapply == "received":
                status = ["sample_received"]
            elif artoapply == "received_tobeverified":
                status = ["sample_received", "attachment_due", "to_be_verified"]

            over = [False, False]
            if override == "nooverride":
                over = [False, False]
            elif override == "override":
                over = [True, False]
            elif override == "overrideempty":
                over = [True, True]

            importer = AnalysisResultsImporter(
                parser=parser,
                context=context,
                allowed_ar_states=status,
                allowed_analysis_states=None,
                override=over,
                instrument_uid=instrument,
            )

            try:
                importer.process()
                errors = importer.errors
                logs = importer.logs
                warns = importer.warns
            except Exception as e:
                errors.extend([repr(e), traceback.format_exc()])

        results = {"errors": errors, "log": logs, "warns": warns}

        return json.dumps(results)
